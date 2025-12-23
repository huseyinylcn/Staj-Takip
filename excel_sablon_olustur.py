"""
Excel şablon dosyası oluşturma scripti
Öğrenci içe aktarma için örnek Excel dosyası oluşturur
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def excel_sablon_olustur():
    """Excel şablon dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Öğrenci Listesi"
    
    # Başlık satırı
    headers = ['Ad', 'Soyad', 'Öğrenci No', 'Telefon']
    ws.append(headers)
    
    # Başlık stilini ayarla
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Örnek veriler ekle
    ornek_veriler = [
        ['Ahmet', 'Yılmaz', '20210101', '05551234567'],
        ['Mehmet', 'Demir', '20210102', '05559876543'],
        ['Ayşe', 'Kaya', '20210103', '05551112233'],
    ]
    
    for veri in ornek_veriler:
        ws.append(veri)
    
    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Açıklama sayfası ekle
    ws2 = wb.create_sheet("Açıklama")
    
    aciklamalar = [
        ["STAJ TAKİP SİSTEMİ - EXCEL ŞABLONU"],
        [""],
        ["KULLANIM:"],
        ["1. 'Öğrenci Listesi' sekmesindeki örnek verileri silin"],
        ["2. Kendi öğrenci verilerinizi ekleyin"],
        ["3. Dosyayı kaydedin"],
        ["4. Staj Takip Sistemi'nde 'Excel Yükle' butonuna tıklayın"],
        ["5. Bu dosyayı seçin"],
        [""],
        ["ZORUNLU ALANLAR:"],
        ["- Ad: Öğrencinin adı"],
        ["- Soyad: Öğrencinin soyadı"],
        ["- Öğrenci No: Benzersiz öğrenci numarası"],
        [""],
        ["OPSİYONEL ALANLAR:"],
        ["- Telefon: İletişim numarası"],
        [""],
        ["DİKKAT:"],
        ["- Öğrenci numaraları benzersiz olmalıdır"],
        ["- İlk satır (başlık satırı) değiştirilmemelidir"],
        ["- Boş satır bırakmayın"],
    ]
    
    for satir in aciklamalar:
        ws2.append(satir)
    
    # Açıklama sayfası formatı
    ws2.column_dimensions['A'].width = 60
    ws2['A1'].font = Font(bold=True, size=14, color="4472C4")
    
    # Kaydet
    dosya_adi = 'ornek_ogrenci_listesi.xlsx'
    wb.save(dosya_adi)
    print(f"✅ Excel şablon dosyası oluşturuldu: {dosya_adi}")

if __name__ == '__main__':
    excel_sablon_olustur()

