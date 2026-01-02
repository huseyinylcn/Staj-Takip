import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from models import Ogrenci, ZiyaretNotu, StajDegerlendirme
from services.staj_service import OgrenciService

class ExcelService:
    """Excel import/export işlemleri servisi"""
    
    @staticmethod
    def ogrencileri_excel_aktar(dosya_yolu='ogrenci_listesi.xlsx', sinif_id=None):
        """Öğrencileri Excel'e aktar (opsiyonel: sınıf filtresi)"""
        ogrenciler = OgrenciService.tum_ogrencileri_getir(sinif_id=sinif_id)
        
        # Workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Öğrenci Listesi"
        
        # Başlık satırı
        headers = ['ID', 'Sınıf', 'Ad', 'Soyad', 'Öğrenci No', 'Telefon', 'Kayıt Tarihi', 
                   'Toplam Ziyaret', 'Staj Notu', 'Harf Notu']
        ws.append(headers)
        
        # Başlık stilini ayarla
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Öğrenci verilerini ekle
        for ogrenci in ogrenciler:
            ziyaret_sayisi = len(ogrenci.ziyaret_notlari)
            
            if ogrenci.staj_degerlendirme:
                toplam = ogrenci.staj_degerlendirme.toplam
                harf = ogrenci.staj_degerlendirme.harf_notu
            else:
                toplam = '-'
                harf = '-'
            
            ws.append([
                ogrenci.id,
                ogrenci.sinif.ad if ogrenci.sinif else '-',
                ogrenci.ad,
                ogrenci.soyad,
                ogrenci.ogrenci_no,
                ogrenci.telefon or '-',
                ogrenci.kayit_tarihi.strftime('%Y-%m-%d') if ogrenci.kayit_tarihi else '-',
                ziyaret_sayisi,
                toplam,
                harf
            ])
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        
        # Kaydet
        wb.save(dosya_yolu)
        return dosya_yolu
    
    @staticmethod
    def ogrencileri_excel_ice_aktar(dosya_yolu):
        """Excel'den öğrenci listesi içe aktar"""
        try:
            df = pd.read_excel(dosya_yolu)
            
            eklenen = 0
            hatalar = []
            
            for index, row in df.iterrows():
                try:
                    ad = str(row.get('Ad', row.get('ad', ''))).strip()
                    soyad = str(row.get('Soyad', row.get('soyad', ''))).strip()
                    ogrenci_no = str(row.get('Öğrenci No', row.get('ogrenci_no', ''))).strip()
                    telefon = str(row.get('Telefon', row.get('telefon', ''))) if pd.notna(row.get('Telefon', row.get('telefon', ''))) else None
                    
                    if not ad or not soyad or not ogrenci_no:
                        hatalar.append(f"Satır {index + 2}: Eksik bilgi")
                        continue
                    
                    OgrenciService.ogrenci_ekle(ad, soyad, ogrenci_no, telefon)
                    eklenen += 1
                    
                except ValueError as e:
                    hatalar.append(f"Satır {index + 2}: {str(e)}")
                except Exception as e:
                    hatalar.append(f"Satır {index + 2}: Beklenmeyen hata - {str(e)}")
            
            return {
                'basarili': True,
                'eklenen': eklenen,
                'hatalar': hatalar
            }
            
        except Exception as e:
            return {
                'basarili': False,
                'hata': str(e)
            }
    
    @staticmethod
    def degerlendirme_raporu_olustur(dosya_yolu='degerlendirme_raporu.xlsx', sinif_id=None):
        """Detaylı değerlendirme raporu oluştur (opsiyonel: sınıf filtresi)"""
        ogrenciler = OgrenciService.tum_ogrencileri_getir(sinif_id=sinif_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Staj Değerlendirme"
        
        # Başlık satırı
        headers = [
            'Sınıf', 'Öğrenci No', 'Ad Soyad', 
            'İşyeren Notu (30p)', 'İçindekiler (10p)', 'Firma Bilgisi (10p)',
            'Yazım Düzeni (10p)', 'Resim/Şekil (10p)', 'Dil Kullanımı (20p)',
            'Sonuç Bölümü (10p)', 'Defter Düzeni/Mülakat (30p)',
            'TOPLAM', 'HARF NOTU'
        ]
        ws.append(headers)
        
        # Başlık stili
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Öğrenci verilerini ekle
        for ogrenci in ogrenciler:
            deg = ogrenci.staj_degerlendirme
            
            if deg:
                ws.append([
                    ogrenci.sinif.ad if ogrenci.sinif else '-',
                    ogrenci.ogrenci_no,
                    f"{ogrenci.ad} {ogrenci.soyad}",
                    deg.isyeren_notu,
                    deg.icindekiler,
                    deg.firma_bilgisi,
                    deg.yazim_duzeni,
                    deg.resim_sekil,
                    deg.dil_kullanimi,
                    deg.sonuc_bolumu,
                    deg.defter_duzeni_mulakat,
                    deg.toplam,
                    deg.harf_notu
                ])
        
        # Sütun genişliklerini ayarla
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        
        wb.save(dosya_yolu)
        return dosya_yolu
    
    @staticmethod
    def normal_donem_raporu_olustur(dosya_yolu='normal_donem_raporu.xlsx', sinif_id=None):
        """Normal dönem notları raporu oluştur (opsiyonel: sınıf filtresi)"""
        from models import NormalDonemDegerlendirme
        
        ogrenciler = OgrenciService.tum_ogrencileri_getir(sinif_id=sinif_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Normal Dönem Notları"
        
        # Başlık satırı
        headers = [
            'Sınıf', 'Öğrenci No', 'Ad Soyad',
            'Vize Notu', 'Vize Ödev', 'Vize Ödev %', 'Vize Toplam',
            'Final Notu', 'Final Ödev', 'Final Ödev %', 'Final Toplam',
            'Devamsızlık Durumu',
            'Bütünleme Notu', 'Büt Ödev', 'Büt Ödev %', 'Bütünleme Toplam',
            'Genel Toplam', 'Harf Notu'
        ]
        ws.append(headers)
        
        # Başlık stili
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Öğrenci verilerini ekle
        for ogrenci in ogrenciler:
            deg = ogrenci.normal_donem_degerlendirme
            
            if deg:
                devamsizlik_durumu = "Geçti" if deg.devamsizlik_durumu else "Kaldı"
                
                ws.append([
                    ogrenci.sinif.ad if ogrenci.sinif else '-',
                    ogrenci.ogrenci_no,
                    f"{ogrenci.ad} {ogrenci.soyad}",
                    deg.vize_notu if deg.vize_notu else 0,
                    deg.vize_odev_puani if deg.vize_odev_puani is not None else '-',
                    deg.vize_odev_yuzdesi if deg.vize_odev_yuzdesi > 0 else '-',
                    round(deg.vize_toplam, 2) if deg.vize_toplam else 0,
                    deg.final_notu if deg.final_notu else 0,
                    deg.final_odev_puani if deg.final_odev_puani is not None else '-',
                    deg.final_odev_yuzdesi if deg.final_odev_yuzdesi > 0 else '-',
                    round(deg.final_toplam, 2) if deg.final_toplam else 0,
                    devamsizlik_durumu,
                    deg.butunleme_notu if deg.butunleme_notu is not None else '-',
                    deg.butunleme_odev_puani if deg.butunleme_odev_puani is not None else '-',
                    deg.butunleme_odev_yuzdesi if deg.butunleme_odev_yuzdesi > 0 else '-',
                    round(deg.butunleme_toplam, 2) if deg.butunleme_toplam is not None else '-',
                    round(deg.genel_toplam, 2) if deg.genel_toplam else 0,
                    deg.harf_notu if deg.harf_notu else '-'
                ])
        
        # Sütun genişliklerini ayarla
        column_widths = {
            'A': 15,  # Sınıf
            'B': 15,  # Öğrenci No
            'C': 20,  # Ad Soyad
            'D': 12,  # Vize Notu
            'E': 12,  # Vize Ödev
            'F': 12,  # Vize Ödev %
            'G': 12,  # Vize Toplam
            'H': 12,  # Final Notu
            'I': 12,  # Final Ödev
            'J': 12,  # Final Ödev %
            'K': 12,  # Final Toplam
            'L': 18,  # Devamsızlık
            'M': 15,  # Bütünleme Notu
            'N': 12,  # Büt Ödev
            'O': 12,  # Büt Ödev %
            'P': 15,  # Bütünleme Toplam
            'Q': 15,  # Genel Toplam
            'R': 12   # Harf Notu
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        wb.save(dosya_yolu)
        return dosya_yolu

